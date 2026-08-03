"""紙本「產品出/進貨抽檢記錄」版面重刻(照片版草稿)。

依現場表單照片(docs/S__43737172.jpg)以 openpyxl 重建版面:
- 標題列 + 表頭區塊(拆櫃/QC 日期、櫃號/封籤號)
- 主表:檢驗項目 | 檢驗規範 | 各型號(每型號兩個子欄,範圍項拆上下限)
- 異常值黃底紅字(對應現場黃螢光筆習慣)
- 底部抽樣標準區(AQL 允收/實際 — 草稿先留空欄位)

註:拿到公司實際 Excel 模板後,此檔將改為「開模板填值」,
   對外介面 build_workbook(sheet) 不變。
"""

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.inspections.judge import _resolve_range
from app.models import InspectionSheet, ModelInspection

THIN = Side(style="thin", color="333333")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
TITLE_FONT = Font(size=14, bold=True)
HEADER_FONT = Font(size=10, bold=True)
BODY_FONT = Font(size=10)
FAIL_FONT = Font(size=10, bold=True, color="C00000")
HIGHLIGHT = PatternFill(start_color="FFF2A8", end_color="FFF2A8", fill_type="solid")
HEADER_FILL = PatternFill(start_color="EFEFEF", end_color="EFEFEF", fill_type="solid")

LABEL_COL = 1  # A:檢驗項目
SPEC_COL = 2  # B:檢驗規範
MODEL_WIDTH = 2  # 每型號佔兩個子欄(範圍項拆上下限)


def _cell(ws: Worksheet, row: int, col: int, value, *, font=BODY_FONT, align=CENTER, fill=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = font
    cell.alignment = align
    cell.border = BORDER
    if fill:
        cell.fill = fill
    return cell


def _merge(ws: Worksheet, row: int, col_start: int, col_end: int, value, **kwargs):
    ws.merge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_end)
    cell = _cell(ws, row, col_start, value, **kwargs)
    # 合併範圍內每格都要框線,否則右/下邊缺線
    for col in range(col_start, col_end + 1):
        ws.cell(row=row, column=col).border = BORDER
    return cell


def _model_cols(index: int) -> tuple[int, int]:
    start = SPEC_COL + 1 + index * MODEL_WIDTH
    return start, start + MODEL_WIDTH - 1


def _is_highlighted(mi: ModelInspection, key: str) -> bool:
    return any(h.get("item") == key for h in (mi.judgement or {}).get("highlights", []))


def _measured_text(mi: ModelInspection, item: dict) -> str:
    key = item["key"]
    if item.get("source_type") == "auto":
        # 自動帶入項:型號欄直接印標準值(同紙本,例:標稱功率 15|15)
        return item.get("standard_text", "—")
    if item.get("source_type") == "pdf":
        values = [
            str(s.photometric.get(key, "—"))
            for s in sorted(mi.samples, key=lambda s: s.seq)
        ]
        return " / ".join(values) if values else "—"
    return str(mi.item_values.get(key, "—"))


def build_workbook(sheet: InspectionSheet) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "抽檢記錄"
    models = sheet.model_inspections
    last_col = SPEC_COL + max(1, len(models)) * MODEL_WIDTH

    # ── 標題 ──
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    title_category = models[0].spec.product_category if models else ""
    cell = ws.cell(row=1, column=1, value=f"產品出/進貨抽檢記錄-{title_category}")
    cell.font = TITLE_FONT
    cell.alignment = CENTER
    ws.row_dimensions[1].height = 26

    # ── 表頭區塊 ──
    mid = max(2, last_col // 2)
    _merge(ws, 2, 1, mid, f"拆櫃日期:{sheet.unstuffing_date or '—'}", align=LEFT)
    _merge(ws, 2, mid + 1, last_col, f"櫃號:{sheet.container_no}", align=LEFT)
    _merge(ws, 3, 1, mid, f"QC日期:{sheet.qc_date or '—'}", align=LEFT)
    _merge(ws, 3, mid + 1, last_col, f"封籤號:{sheet.seal_no or '—'}", align=LEFT)

    # ── 主表表頭 ──
    row = 5
    _cell(ws, row, LABEL_COL, "檢驗項目", font=HEADER_FONT, fill=HEADER_FILL)
    _cell(ws, row, SPEC_COL, "檢驗規範", font=HEADER_FONT, fill=HEADER_FILL)
    for i, mi in enumerate(models):
        c1, c2 = _model_cols(i)
        _merge(ws, row, c1, c2, mi.product_name, font=HEADER_FONT, fill=HEADER_FILL)
    row += 1

    if models:
        _cell(ws, row, LABEL_COL, "批號", fill=HEADER_FILL)
        _cell(ws, row, SPEC_COL, "燈體", fill=HEADER_FILL)
        for i, mi in enumerate(models):
            c1, c2 = _model_cols(i)
            _merge(ws, row, c1, c2, mi.batch_code or "—")
        row += 1

    # ── 檢驗項目列(以第一個型號的標準模板為列骨架)──
    template_items = (models[0].spec.items or []) if models else []
    for item in template_items:
        key = item["key"]
        rule = item.get("rule", {})
        lo, hi = _resolve_range(rule) if rule.get("type") == "range" else (None, None)
        has_bounds = lo is not None and hi is not None

        # 範圍項:第一行為各型號的上下限拆格
        if has_bounds:
            _cell(ws, row, LABEL_COL, item.get("label", key))
            _cell(ws, row, SPEC_COL, item.get("standard_text", ""))
            for i, _mi in enumerate(models):
                c1, c2 = _model_cols(i)
                _cell(ws, row, c1, f"{lo:g}")
                _cell(ws, row, c2, f"{hi:g}")
            row += 1
            label_cell, spec_cell = "", ""  # 實測行標籤留空(同紙本)
        else:
            label_cell = item.get("label", key)
            spec_cell = item.get("standard_text", "")

        # 實測/勾選行
        _cell(ws, row, LABEL_COL, label_cell)
        _cell(ws, row, SPEC_COL, spec_cell)
        for i, mi in enumerate(models):
            c1, c2 = _model_cols(i)
            failed = _is_highlighted(mi, key)
            _merge(
                ws, row, c1, c2, _measured_text(mi, item),
                font=FAIL_FONT if failed else BODY_FONT,
                fill=HIGHLIGHT if failed else None,
            )
        row += 1

    # ── 底部抽樣標準區(AQL — 草稿:欄位結構先到位,數據待接)──
    row += 1
    _merge(ws, row, 1, SPEC_COL, "抽樣標準", font=HEADER_FONT, fill=HEADER_FILL)
    for i, _mi in enumerate(models):
        c1, c2 = _model_cols(i)
        _cell(ws, row, c1, "允收數量", font=HEADER_FONT, fill=HEADER_FILL)
        _cell(ws, row, c2, "實際數量", font=HEADER_FONT, fill=HEADER_FILL)
    row += 1
    for aql_label in ("嚴重缺失 0.65", "主要缺失 1.0", "次要缺失 1.5"):
        _merge(ws, row, 1, SPEC_COL, aql_label, align=LEFT)
        for i, _mi in enumerate(models):
            c1, c2 = _model_cols(i)
            _cell(ws, row, c1, "")
            _cell(ws, row, c2, "")
        row += 1

    # ── 欄寬 ──
    ws.column_dimensions[get_column_letter(LABEL_COL)].width = 24
    ws.column_dimensions[get_column_letter(SPEC_COL)].width = 26
    for col in range(SPEC_COL + 1, last_col + 1):
        ws.column_dimensions[get_column_letter(col)].width = 13

    # 列印設定:A4 直式塞一頁寬(同紙本)
    ws.print_options.horizontalCentered = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    return wb
