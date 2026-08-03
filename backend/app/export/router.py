from io import BytesIO

from fastapi import APIRouter, Depends, HTTPException, status
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy.orm import Session

from app.auth.service import get_current_user
from app.database import get_db
from app.models import InspectionSheet, User

router = APIRouter(prefix="/api/export", tags=["export"])

HIGHLIGHT = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
FAIL_FONT = Font(color="C00000", bold=True)
HEADER_FONT = Font(bold=True)
CENTER = Alignment(horizontal="center", vertical="center")


@router.get("/sheets/{sheet_id}/xlsx")
def export_xlsx(
    sheet_id: int, db: Session = Depends(get_db), _user: User = Depends(get_current_user)
) -> StreamingResponse:
    """匯出檢驗報表 Excel。

    注意:目前為系統預設版面;待取得公司現行報表模板檔後改為模板填值,
    使格式與紙本 100% 一致(docs/requirements.md §7)。
    """
    sheet = db.get(InspectionSheet, sheet_id)
    if sheet is None:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "檢驗單不存在")

    wb = Workbook()
    ws = wb.active
    ws.title = "抽檢記錄"

    ws["A1"] = "產品出/進貨抽檢記錄"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"櫃號:{sheet.container_no}"
    ws["B2"] = f"封籤號:{sheet.seal_no}"
    ws["C2"] = f"拆櫃日期:{sheet.unstuffing_date}"
    ws["D2"] = f"QC日期:{sheet.qc_date}"
    ws["E2"] = f"狀態:{sheet.status}"

    row = 4
    ws.cell(row=row, column=1, value="檢驗項目").font = HEADER_FONT
    ws.cell(row=row, column=2, value="檢驗規範").font = HEADER_FONT
    col = 3
    for mi in sheet.model_inspections:
        cell = ws.cell(row=row, column=col, value=f"{mi.product_name}({mi.result})")
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        col += 1
    row += 1

    # 逐項輸出:以第一個型號的標準模板為列骨架(同單各型號模板通常一致)
    if sheet.model_inspections:
        template_items = sheet.model_inspections[0].spec.items or []
        for item in template_items:
            key = item["key"]
            ws.cell(row=row, column=1, value=item.get("label", key))
            ws.cell(row=row, column=2, value=item.get("standard_text", ""))
            col = 3
            for mi in sheet.model_inspections:
                if item.get("source_type") == "pdf":
                    values = [
                        f"#{s.seq}: {s.photometric.get(key, '—')}"
                        for s in sorted(mi.samples, key=lambda s: s.seq)
                    ]
                    text = " / ".join(values) if values else "—"
                else:
                    text = str(mi.item_values.get(key, "—"))
                cell = ws.cell(row=row, column=col, value=text)
                # 異常高亮(取代人工黃螢光筆)
                highlights = (mi.judgement or {}).get("highlights", [])
                if any(h.get("item") == key for h in highlights):
                    cell.fill = HIGHLIGHT
                    cell.font = FAIL_FONT
                col += 1
            row += 1

    for column_cells in ws.columns:
        width = max(len(str(c.value or "")) for c in column_cells) + 2
        ws.column_dimensions[column_cells[0].column_letter].width = min(width, 40)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    filename = f"inspection_{sheet.container_no}_{sheet_id}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
